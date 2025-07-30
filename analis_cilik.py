import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
from datetime import datetime
import os
from openpyxl import load_workbook
from fpdf import FPDF

st.set_page_config(page_title="Adab Berpakaian", layout="centered", page_icon="👕")

st.markdown("""
<style>
    .main {
        background-color: #f0f9ff;
    }
    .stApp {
        font-family: 'Segoe UI', sans-serif;
    }
</style>
""", unsafe_allow_html=True)

st.title("👕 Analisis Adab Berpakaian")
st.markdown("Yuk isi hasil survei teman dan lihat hasil analisismu!")

# Input user (siswa)
st.sidebar.title("🧒 Identitas Kelompok")
nama_user = st.sidebar.text_input("Nama Kamu")

# Path file Excel
folder = "data_output"
os.makedirs(folder, exist_ok=True)
filename = os.path.join(folder, "hasil_survei_siswa.xlsx")

if nama_user:
    st.success(f"Data akan disimpan atas nama: **{nama_user}**")

    # Form Input Data
    with st.form("form_input"):
        st.subheader("📝 Masukkan Data Wawancara Teman")
        nama = st.text_input("Nama Teman")
        warna = st.selectbox("Warna Pakaian", ["Putih", "Hitam", "Biru", "Merah", "Lainnya"])
        jenis = st.text_input("Jenis Pakaian")
        aurat = st.radio("Apakah sudah menutup aurat?", ["✓", "✗"])
        submitted = st.form_submit_button("✅ Tambahkan Data")

        if submitted:
            st.success(f"Data untuk {nama} berhasil ditambahkan!")
            new_data = {
                "Nama User": nama_user,
                "Nama Teman": nama,
                "Warna": warna,
                "Jenis Pakaian": jenis,
                "Menutup Aurat": aurat
            }

            # Simpan ke Excel bersama
            df_new = pd.DataFrame([new_data])
            if os.path.exists(filename):
                df_existing = pd.read_excel(filename)
                df_combined = pd.concat([df_existing, df_new], ignore_index=True)
            else:
                df_combined = df_new

            df_combined.to_excel(filename, index=False)
            st.info(f"📁 Data ditambahkan ke file: `{filename}`")

    # Baca dan tampilkan data user
    if os.path.exists(filename):
        df_all = pd.read_excel(filename)
        df = df_all[df_all["Nama User"] == nama_user]
    else:
        df = pd.DataFrame()

    if not df.empty:
        st.subheader(f"📋 Data Hasil Wawancara oleh {nama_user}")
        st.dataframe(df, use_container_width=True)

        # ======== CRUD FITUR ========
        st.subheader("🛠️ Edit / Hapus Data")
        pilihan_index = st.selectbox("Pilih data yang ingin diedit / dihapus:", df.index, format_func=lambda x: f"{df.loc[x, 'Nama Teman']} - {df.loc[x, 'Warna']}")

        with st.form("form_edit"):
            nama_edit = st.text_input("Edit Nama Teman", df.loc[pilihan_index, 'Nama Teman'])
            warna_edit = st.selectbox("Edit Warna Pakaian", ["Putih", "Hitam", "Biru", "Merah", "Lainnya"], index=["Putih", "Hitam", "Biru", "Merah", "Lainnya"].index(df.loc[pilihan_index, 'Warna']))
            jenis_edit = st.text_input("Edit Jenis Pakaian", df.loc[pilihan_index, 'Jenis Pakaian'])
            aurat_edit = st.radio("Edit Status Aurat", ["✓", "✗"], index=["✓", "✗"].index(df.loc[pilihan_index, 'Menutup Aurat']))

            col1, col2 = st.columns(2)
            with col1:
                update_btn = st.form_submit_button("📏 Simpan Perubahan")
            with col2:
                delete_btn = st.form_submit_button("🖑️ Hapus Data")

            if update_btn:
                index_global = df_all[df_all["Nama User"] == nama_user].index[pilihan_index]
                df_all.loc[index_global, ['Nama Teman', 'Warna', 'Jenis Pakaian', 'Menutup Aurat']] = [
                    nama_edit, warna_edit, jenis_edit, aurat_edit
                ]
                df_all.to_excel(filename, index=False)
                st.success("✅ Data berhasil diperbarui!")
                st.experimental_rerun()

            if delete_btn:
                df_all.drop(index=df.index[pilihan_index], inplace=True)
                df_all.to_excel(filename, index=False)
                st.success("🗑️ Data berhasil dihapus.")
                st.session_state.clear()
                st.experimental_rerun()


        # ======== Visualisasi dan PDF ========
        st.subheader("📊 Grafik Warna Pakaian")
        warna_count = df['Warna'].value_counts()
        fig, ax = plt.subplots()
        warna_count.plot(kind='bar', color=['#5fa9f0', '#ff7676', '#f0c05f', '#7c83fd', '#4dc9a6'], ax=ax)
        ax.set_xlabel("Warna Pakaian")
        ax.set_ylabel("Jumlah Teman")
        ax.set_title("Jumlah Teman Berdasarkan Warna Pakaian")
        st.pyplot(fig)

        st.subheader("🦕 Grafik Menutup Aurat")
        aurat_count = df['Menutup Aurat'].value_counts()
        fig2, ax2 = plt.subplots()
        ax2.pie(aurat_count, labels=aurat_count.index, autopct='%1.1f%%', colors=['#86efac', '#fda4af'])
        ax2.set_title("Persentase Teman yang Menutup Aurat")
        st.pyplot(fig2)

        st.subheader("🧠 Kesimpulan Otomatis")
        total = len(df)
        warna_terbanyak = warna_count.idxmax() if not warna_count.empty else "-"
        aurat_ok = aurat_count.get("✓", 0)
        kesimpulan = f"Dari {total} responden, warna pakaian terbanyak adalah {warna_terbanyak}. Sebanyak {aurat_ok} dari {total} teman memakai pakaian yang menutup aurat."
        st.info(kesimpulan)

        if aurat_ok == total and total > 0:
            st.balloons()
            st.success("👍 Semua teman sudah menutup aurat, keren banget!")
        elif aurat_ok > total // 2:
            st.success("👌 Mayoritas teman sudah berpakaian sesuai adab.")
        else:
            st.warning("⚠️ Masih banyak teman yang perlu belajar adab berpakaian.")

        # PDF Download
        st.subheader("📄 Unduh Hasil Analisis (PDF)")
        pdf_file = os.path.join(folder, f"{nama_user.replace(' ', '_')}_laporan.pdf")
        warna_chart = os.path.join(folder, f"{nama_user}_warna.png")
        aurat_chart = os.path.join(folder, f"{nama_user}_aurat.png")

        def create_pdf(dataframe, kesimpulan_text, pdf_path, warna_chart_path, aurat_chart_path):
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", 'B', 14)
            pdf.cell(0, 10, "Laporan Survei Adab Berpakaian", ln=True, align="C")
            pdf.ln(5)
            pdf.set_font("Arial", '', 12)
            pdf.cell(0, 10, f"Nama Siswa: {nama_user}", ln=True)
            pdf.cell(0, 10, f"Tanggal: {datetime.now().strftime('%d-%m-%Y')}", ln=True)
            pdf.ln(5)

            pdf.set_font("Arial", 'B', 12)
            pdf.cell(0, 10, "Data Teman yang Disurvei:", ln=True)
            pdf.set_font("Arial", '', 11)
            for idx, row in dataframe.iterrows():
                aurat_status = "Sudah" if row['Menutup Aurat'] == '✓' else "Belum"
                line = f"{idx+1}. {row['Nama Teman']} - {row['Warna']} - {row['Jenis Pakaian']} - Aurat: {aurat_status}"
                pdf.multi_cell(0, 8, txt=line)
            pdf.ln(5)

            if os.path.exists(warna_chart_path):
                pdf.set_font("Arial", 'B', 12)
                pdf.cell(0, 10, "Grafik Warna Pakaian:", ln=True)
                pdf.image(warna_chart_path, w=180)
                pdf.ln(5)

            if os.path.exists(aurat_chart_path):
                pdf.set_font("Arial", 'B', 12)
                pdf.cell(0, 10, "Grafik Menutup Aurat:", ln=True)
                pdf.image(aurat_chart_path, w=180)
                pdf.ln(5)

            pdf.set_font("Arial", 'B', 12)
            pdf.cell(0, 10, "Kesimpulan:", ln=True)
            pdf.set_font("Arial", '', 11)
            pdf.multi_cell(0, 8, txt=kesimpulan_text)
            pdf.output(pdf_path)

        if st.button("📅 Download PDF"):
            fig.savefig(warna_chart)
            fig2.savefig(aurat_chart)
            create_pdf(df, kesimpulan, pdf_file, warna_chart, aurat_chart)
            with open(pdf_file, "rb") as f:
                st.download_button("Klik untuk mengunduh", f, file_name=os.path.basename(pdf_file), mime="application/pdf")
            if os.path.exists(warna_chart): os.remove(warna_chart)
            if os.path.exists(aurat_chart): os.remove(aurat_chart)

    else:
        st.info("Belum ada data yang kamu masukkan. Silakan isi formulir di atas.")

else:
    st.warning("Silakan isi nama kamu di sidebar terlebih dahulu.")
